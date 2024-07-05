import tkinter as tk
from tkinter import ttk, simpledialog, messagebox
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl as opx
from ttkthemes import ThemedStyle
# Importation des données
df = pd.read_csv("D:\\project-Bi\\université\\classement.csv", sep=";")

# Fonction pour sauvegarder le DataFrame au format Excel
def save_excel(df, nameF, sheetname):
    writer = pd.ExcelWriter(nameF, engine='xlsxwriter')
    df.to_excel(writer, sheet_name=sheetname, index=False)
    writer.save()

# Fonction pour sauvegarder le plot au format Excel
def save_plot(plt, nameF, sheetname):
    if not nameF.endswith('.xlsx'):
        nameF += '.xlsx'
    plt.figure(figsize=(8, 5))
    plt.plot(df['year'], df['country'])
    plt.savefig('scatter_plot.png')
    plt.close()

    try:
        workbook = opx.load_workbook(nameF)
    except FileNotFoundError:
        workbook = opx.Workbook()
    if sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]
    else:
        worksheet = workbook.create_sheet(sheetname)
    img = opx.drawing.image.Image('scatter_plot.png')
    worksheet.add_image(img)
    workbook.save(nameF)

# Fonction pour le top 10 des universités en brevets pour une année donnée
def top10_patents(df, year):
    if 2017 <= year <= 2020:
        df_year = df[df['year'] == year]
        top_universities = df_year.sort_values(by='patents', ascending=False).head(10)
        return top_universities[['patents', 'country', 'institution']]
    else:
        messagebox.showerror("Erreur", "L'année doit être entre 2017 et 2020.")
        return pd.DataFrame()  # Retourner un DataFrame vide ou une autre valeur en cas d'erreur

def display_table(data):
    root = tk.Tk()
    root.title("Top 10 Universities for Patents")
    root.geometry("700x200")
    root.configure(bg="lightgray")
    # Créer un tableau (Treeview) pour afficher les données
    tree = ttk.Treeview(root)
    tree["columns"] = tuple(data.columns)
    tree["show"] = "headings"
    # Utiliser un thème ttkthemes pour avoir un meilleur contrôle sur les couleurs
    style = ThemedStyle(root)
    style.set_theme("plastik")

    # Configurer les colonnes du tableau
    for col in data.columns:
        tree.heading(col, text=col, anchor="center")
        tree.column(col, width=100, anchor="center")

    # Ajouter les données au tableau
    for index, row in data.iterrows():
        tree.insert("", "end", values=tuple(row))

    # Afficher le tableau
    tree.pack(expand=True, fill="both")

    root.mainloop()

# Fonction pour obtenir la première et la dernière université pour une année donnée
def get_First_Last_University(df):
    while True:
        annee = simpledialog.askinteger("Année", "Taper l'année:")
        if 2017 <= annee <= 2020:
            break
        else:
            print('Entrer une année entre 2017 et 2020')

    df_annee = df[df['year'] == annee]
    df_annee = df_annee.sort_values(by='world_rank')
    first_university = df_annee.iloc[0]['institution']
    last_university = df_annee.iloc[-1]['institution']
    result = f'First University: {first_university}\nLast University: {last_university}'
    show_result_window(result)

# Fonction pour les scores pour une année donnée
def score(df, year):
    df_year = df[df['year'] == year]
    top_scores = df_year.sort_values(by='score', ascending=False).head(10)
    min_scores = df_year.sort_values(by='score').head(10)
    top_scores_grouped = top_scores.groupby('country').head(1)[['country', 'institution', 'score']]
    min_scores_grouped = min_scores.groupby('country').head(1)[['country', 'institution', 'score']]
    result = f'Top 10 Scores:\n{top_scores_grouped}\n\nMin 10 Scores:\n{min_scores_grouped}'
    show_result_window(result)

# Fonction pour le total des universités classées par pays pour une année donnée
def total_ranked_universities(df):
    while True:
        annee = simpledialog.askinteger("Année", "Taper l'année:")
        if 2017 <= annee <= 2020:
            break
        else:
            print('Entrer une année entre 2017 et 2020')

    result_df = df[df['year'] == annee]
    total_by_country = result_df['country'].value_counts().sort_values(ascending=False)
    if total_by_country.empty:
        result = "No data available for the specified year."
    else:
        result_df = total_by_country.reset_index()
        result_df.columns = ['country', 'total_universities']
        result = f'Total Universities Ranked by Country for {annee}:\n{result_df}'
    show_result_window(result)

# Fonction pour les universités avec le classement mondial maximum par année
def max_universities(df):
    new_df = df[['year', 'country', 'world_rank']]
    maximum = new_df.groupby('year')['world_rank'].idxmax()
    new_df = new_df.loc[maximum]
    result = f'Universities with Maximum World Rank by Year:\n{new_df}'
    show_result_window(result)

# Fonction pour créer un graphique à barres pour les 10 premiers pays avec le total d'universités
def create_bar_graphic(df):
    result_df = total_ranked_universities(df)
    if result_df is None or result_df.empty:
        result = "No data available for the specified year."
    else:
        top_10_countries = result_df.head(10)
        plt.figure(figsize=(8, 5))
        plt.bar(top_10_countries['country'], top_10_countries['total_universities'], color='blue')
        plt.xlabel('Pays', color='blue')
        plt.ylabel('Nombre d Universities', color='green')
        plt.title(f'Nombre d Université par pays ({year})', color='red')
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plot_filename = 'plot.png'
        plt.savefig(plot_filename)
        nameF = input("Nom du fichier de plot Excel: ")
        sheetname = input("Nom de la feuille Excel: ")
        save_plot(plt, nameF, sheetname)
        result = f'Bar Graphic Created and Saved as {plot_filename} in {nameF} under {sheetname} sheet.'
    show_result_window(result)

# Fonction pour créer un graphique de dispersion pour les 5 meilleures universités d'un pays pour une année donnée
def create_scatter_graphic(df):
    while True:
        annee = simpledialog.askinteger("Année", "Taper l'année:")
        if 2017 <= annee <= 2020:
            break
        else:
            print('Entrer une année entre 2017 et 2020')
    country = simpledialog.askstring("Pays", "Veuillez saisir le nom du pays:")
    top5_universities = df[(df['year'] == annee) & (df['country'] == country)].nlargest(5, 'score')
    plt.figure(figsize=(8, 5))
    plt.scatter(top5_universities['score'], top5_universities['institution'])
    plt.xlabel('Score', color='blue')
    plt.ylabel('Institution', color='green')
    plt.title(f'Top 5 universités pour {country} en {annee}', color='red')
    plt.tight_layout()
    plot_filename = 'scatter_plot.png'
    plt.savefig(plot_filename)
    nameF = input("Nom du fichier de plot Excel: ")
    sheetname = input("Nom de la feuille Excel: ")
    save_plot(plt, nameF, sheetname)
    result = f'Scatter Graphic Created and Saved as {plot_filename} in {nameF} under {sheetname} sheet.'
    show_result_window(result)

# Fonction pour afficher les résultats
def show_result_window(result):
    result_window = tk.Toplevel(window)
    result_window.title("Résultat")
    result_window.geometry("400x300")
    result_window.configure(bg="lightblue")
    result_label = tk.Label(result_window, text=result)
    result_label.pack(pady=50)

# Fonction à exécuter lorsque le bouton est cliqué
def on_button_click():
    selected_option = combobox.get()
    
    if selected_option == "Sauvegarder le DataFrame au format Excel":
        nameF = simpledialog.askstring("Nom du fichier Excel", "Nom du fichier Excel : ")
        sheetname = simpledialog.askstring("Nom de la feuille Excel", "Nom de la feuille Excel : ")
        save_excel(df, nameF, sheetname)
        result = f'DataFrame sauvegardé dans {nameF} sous le nom de la feuille {sheetname}'
        show_result_window(result)

    elif selected_option == "Sauvegarder le plot au format Excel":
        nameF = input("Nom du fichier de plot Excel : ")
        sheetname = input("Nom de la feuille Excel : ")
        create_scatter_graphic(df, nameF, sheetname)
        print("Le plot est sauvegardé !")
        show_result_window(result)

    elif selected_option == "Obtenir la première et la dernière université pour une année donnée":
        get_First_Last_University(df)

    elif selected_option == "Top 10 des universités en brevets pour une année donnée":
        year_value = simpledialog.askinteger("Année", "Entrez l'année : ")
        if year_value is not None:
            top10_df = top10_patents(df, year_value)
            display_table(top10_df)
        else:
            result = "Veuillez saisir une année valide."
            show_result_window(result)

    elif selected_option == "Scores pour une année donnée":
        year_value = simpledialog.askinteger("Année", "Entrez l'année : ")
        if year_value is not None:
            score(df, year_value)
        else:
            result = "Veuillez saisir une année valide."
            show_result_window(result)

    elif selected_option == "Total des universités classées par pays pour une année donnée":
        total_ranked_universities(df)

    elif selected_option == "Universités avec le classement mondial maximum par année":
        max_universities(df)

    elif selected_option == "Créer un graphique à barres pour les 10 premiers pays avec le total d'universités":
        create_bar_graphic(df)

    elif selected_option == "Créer un graphique de dispersion pour les 5 meilleures universités d'un pays":
        create_scatter_graphic(df)

    elif selected_option == "Quitter":
        print("Au revoir !")
        window.destroy()  # Fermer la fenêtre principale

# Fonction à exécuter lorsque l'option de la ComboBox est sélectionnée
def on_combobox_selected(event):
    selected_option = combobox.get()
    print(f"Option sélectionnée : {selected_option}")

# Création de la fenêtre principale
window = tk.Tk()
window.title("Application Université")
window.geometry("600x500")
window.configure(bg="lightblue")

# Ajout d'une étiquette (label)
titre = tk.Label(window, text="Classement des universités", bg="lightblue", font=("Arial", 14, "bold"), foreground="white")
titre.place(x=180, y=80)

# Options du menu
menu_options = [
    "Sauvegarder le DataFrame au format Excel",
    "Sauvegarder le plot au format Excel",
    "Top 10 des universités en brevets pour une année donnée",
    "Obtenir la première et la dernière université pour une année donnée",
    "Scores pour une année donnée",
    "Total des universités classées par pays pour une année donnée",
    "Universités avec le classement mondial maximum par année",
    "Créer un graphique à barres pour les 10 premiers pays avec le total d'universités",
    "Créer un graphique de dispersion pour les 5 meilleures universités d'un pays",
    "Quitter"
]

# Créer une Combobox avec les options du menu
combobox = ttk.Combobox(window, values=menu_options, width=80)
combobox.place(x=80, y=115)
combobox.set("Sélectionnez un menu")
# Définir une fonction à appeler lorsqu'une option est sélectionnée
combobox.bind("<<ComboboxSelected>>", on_combobox_selected)

# Ajout d'un bouton
button = tk.Button(window, text="Sélectionner", command=on_button_click, foreground="white", width=30)
button.place(x=200, y=150)

# Lancement de la boucle principale de l'interface graphique
window.mainloop()
